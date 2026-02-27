# -*- coding: utf-8 -*-
"""
Segmentacao Homogenea de Rodovias — App Unificado
Metodos: CDA (CUSUM/Zi), SHS (Spatial Heterogeneity) e MCV (Minimize CV)
Dashboard interativo em Streamlit com Plotly.
"""

import streamlit as st
import pandas as pd
import numpy as np
from scipy.signal import find_peaks
import plotly.graph_objects as go
import plotly.express as px
from io import BytesIO
from datetime import datetime

from homogeneous_segmentation import (
    segment_ids_to_maximize_spatial_heterogeneity,
    segment_ids_to_minimize_coefficient_of_variation,
)

# ======================================================================
#  CONFIGURACAO DA PAGINA
# ======================================================================
st.set_page_config(
    page_title="Segmentacao Unificada — CDA · SHS · MCV",
    page_icon="\U0001f6e3\ufe0f",
    layout="wide",
    initial_sidebar_state="expanded",
)

# -- CSS customizado --
st.markdown("""
<style>
    .main .block-container { padding-top: 1.5rem; }
    .kpi-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 20px; border-radius: 14px; text-align: center;
        color: white; box-shadow: 0 4px 15px rgba(102,126,234,0.3);
        transition: transform 0.2s;
    }
    .kpi-card:hover { transform: translateY(-3px); }
    .kpi-value { font-size: 2rem; font-weight: 800; line-height: 1.2; }
    .kpi-label { font-size: 0.78rem; opacity: 0.85; margin-top: 4px; }
    .kpi-green { background: linear-gradient(135deg, #2ecc71 0%, #27ae60 100%);
                 box-shadow: 0 4px 15px rgba(46,204,113,0.3); }
    .kpi-red   { background: linear-gradient(135deg, #e74c3c 0%, #c0392b 100%);
                 box-shadow: 0 4px 15px rgba(231,76,60,0.3); }
    .kpi-blue  { background: linear-gradient(135deg, #3498db 0%, #2980b9 100%);
                 box-shadow: 0 4px 15px rgba(52,152,219,0.3); }
    .kpi-purple{ background: linear-gradient(135deg, #9b59b6 0%, #8e44ad 100%);
                 box-shadow: 0 4px 15px rgba(155,89,182,0.3); }
    .kpi-orange{ background: linear-gradient(135deg, #f39c12 0%, #e67e22 100%);
                 box-shadow: 0 4px 15px rgba(243,156,18,0.3); }
    .kpi-teal  { background: linear-gradient(135deg, #1abc9c 0%, #16a085 100%);
                 box-shadow: 0 4px 15px rgba(26,188,156,0.3); }
    .story-box {
        background: #f8f9fa; border-left: 5px solid #667eea;
        padding: 15px 20px; border-radius: 0 10px 10px 0;
        margin: 12px 0; font-size: 0.92rem; color: #333;
    }
    [data-testid="stSidebar"] { background: #f0f2f6; }
    .stTabs [data-baseweb="tab"] { font-weight: 600; }
    .dataframe { font-size: 0.85rem !important; }
    hr { border-top: 2px solid #667eea30; margin: 1.5rem 0; }
</style>
""", unsafe_allow_html=True)

# ======================================================================
#  FUNCOES AUXILIARES
# ======================================================================

def kpi_card(value, label, css_class=""):
    return f"""
    <div class="kpi-card {css_class}">
        <div class="kpi-value">{value}</div>
        <div class="kpi-label">{label}</div>
    </div>"""


def story(text):
    st.markdown(f'<div class="story-box">{text}</div>', unsafe_allow_html=True)


# ── CDA helper: fundir segmentos curtos ──
def merge_short_segments(limites, df_local, col_est, min_m):
    limites = list(limites)
    while True:
        lengths = [
            df_local.loc[limites[i+1], col_est] - df_local.loc[limites[i], col_est]
            for i in range(len(limites) - 1)
        ]
        if min(lengths) >= min_m or len(limites) <= 2:
            break
        short_idx = int(np.argmin(lengths))
        if short_idx == 0:
            limites.pop(1)
        elif short_idx == len(lengths) - 1:
            limites.pop(-2)
        else:
            left_combo  = lengths[short_idx - 1] + lengths[short_idx]
            right_combo = lengths[short_idx] + lengths[short_idx + 1]
            if left_combo <= right_combo:
                limites.pop(short_idx)
            else:
                limites.pop(short_idx + 1)
    return np.array(limites)


# ── CDA helper: subdividir segmentos longos ──
def split_long_segments(limites, df_local, col_est, max_m):
    new_limites = [limites[0]]
    for i in range(len(limites) - 1):
        idx_ini = limites[i]; idx_fim = limites[i + 1]
        est_ini = df_local.loc[idx_ini, col_est]
        est_fim = df_local.loc[idx_fim, col_est]
        seg_len = est_fim - est_ini
        if seg_len > max_m:
            n_parts = int(np.ceil(seg_len / max_m))
            target_len = seg_len / n_parts
            subset = df_local.loc[idx_ini:idx_fim]
            for p in range(1, n_parts):
                target_est = est_ini + p * target_len
                closest_idx = (subset[col_est] - target_est).abs().idxmin()
                if closest_idx not in new_limites:
                    new_limites.append(closest_idx)
        new_limites.append(idx_fim)
    return np.array(sorted(set(new_limites)))


# ── SHS/MCV helper: tabela de segmentos a partir de IDs ──
def build_segments_table(df, seg_ids_col, col_est, col_defl):
    seg_groups = df.groupby(seg_ids_col)
    segmentos = []
    for seg_id, grp in seg_groups:
        est_ini = grp[col_est].iloc[0]
        est_fim = grp[col_est].iloc[-1]
        comp = est_fim - est_ini
        defl_media = grp[col_defl].mean()
        defl_std = grp[col_defl].std(ddof=1) if len(grp) > 1 else 0.0
        cv = (defl_std / defl_media * 100) if defl_media != 0 else 0.0
        segmentos.append({
            'SH': int(seg_id),
            'Inicio (m)': est_ini,
            'Fim (m)': est_fim,
            'Comprimento (m)': comp,
            'Defl. Media': round(defl_media, 1),
            'Desvio Padrao': round(defl_std, 1),
            'CV (%)': round(cv, 1),
            'Defl. Caract.': round(defl_media + defl_std, 1),
        })
    return pd.DataFrame(segmentos)


# Paleta hex para segmentos
SEG_COLORS_HEX = [
    '#66c2a5', '#fc8d62', '#8da0cb', '#e78ac3', '#a6d854',
    '#ffd92f', '#e5c494', '#b3b3b3', '#1b9e77', '#d95f02',
    '#7570b3', '#e7298a', '#66a61e', '#e6ab02', '#a6761d',
]

def rgba_from_hex(hex_color, alpha=0.3):
    h = hex_color.lstrip('#')
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return f'rgba({r},{g},{b},{alpha})'


COLOR_PRIMARY = '#667eea'
COLOR_SEC     = '#764ba2'
COLOR_RED     = '#e74c3c'
COLOR_GREEN   = '#27ae60'
COLOR_YELLOW  = '#f39c12'
COLOR_TEAL    = '#1abc9c'


# ── Excel unificado ──
def gerar_excel_unificado(df_data, seg_cda_df, seg_shs_df, seg_mcv_df,
                          col_est, col_defl,
                          picos_arr=None, vales_arr=None, limites_idx_arr=None,
                          col_zi_name=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference, LineChart
    from openpyxl.chart.marker import Marker
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule

    COR_H = 'FF4A6FA5'; COR_H2 = 'FF667EEA'; COR_AC = 'FF764BA2'
    COR_LB = 'FFF2F4F8'; COR_W = 'FFFFFFFF'; COR_R = 'FFE74C3C'; COR_G = 'FF27AE60'

    ft  = Font(name='Calibri', bold=True, size=16, color='FFFFFF')
    fh  = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    fb  = Font(name='Calibri', size=11)
    fH  = PatternFill('solid', fgColor=COR_H)
    fH2 = PatternFill('solid', fgColor=COR_H2)
    fAC = PatternFill('solid', fgColor=COR_AC)
    fL  = PatternFill('solid', fgColor=COR_LB)
    fW  = PatternFill('solid', fgColor=COR_W)
    ac  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    bd  = Border(*[Side(style='thin', color='D0D0D0')]*4)

    wb = Workbook()

    def write_seg_sheet(ws, title, seg_df, fill_h):
        ws.merge_cells('A1:H2')
        ws['A1'].value = title; ws['A1'].font = ft; ws['A1'].fill = fH; ws['A1'].alignment = ac
        cols = list(seg_df.columns)
        for j, cn in enumerate(cols, 1):
            c = ws.cell(row=4, column=j, value=cn)
            c.font = fh; c.fill = fill_h; c.alignment = ac; c.border = bd
            ws.column_dimensions[get_column_letter(j)].width = 18
        for i, rd in seg_df.iterrows():
            for j, cn in enumerate(cols, 1):
                c = ws.cell(row=5+i, column=j, value=rd[cn])
                c.font = fb; c.alignment = ac; c.border = bd
                c.fill = fL if i % 2 == 0 else fW
                if cn in ['Inicio (m)', 'Fim (m)', 'Comprimento (m)']: c.number_format = '#,##0'
                elif cn in ['Defl. Media', 'Desvio Padrao', 'CV (%)', 'Defl. Caract.', 'Zi_fim']:
                    c.number_format = '0.0'
        lr = 5 + len(seg_df) - 1
        if 'Comprimento (m)' in cols:
            comp_col = cols.index('Comprimento (m)') + 1
            ws.conditional_formatting.add(
                f'{get_column_letter(comp_col)}5:{get_column_letter(comp_col)}{lr}',
                DataBarRule(start_type='min', end_type='max', color=COR_H2[2:]))
        if 'Defl. Media' in cols:
            defl_col = cols.index('Defl. Media') + 1
            ch = BarChart(); ch.type='col'; ch.style=10
            ch.title='Deflexao Media por SH'; ch.y_axis.title='Deflexao (0,01mm)'
            ch.width=22; ch.height=12
            ch.add_data(Reference(ws, min_col=defl_col, min_row=4, max_row=lr), titles_from_data=True)
            ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=lr))
            ch.series[0].graphicalProperties.solidFill = COR_H2[2:]
            ch.series[0].dLbls = DataLabelList()
            ch.series[0].dLbls.showVal = True; ch.series[0].dLbls.numFmt = '0.0'
            ch.legend = None
            ws.add_chart(ch, f'A{lr+3}')

    # ── Aba CDA ──
    ws_cda = wb.active; ws_cda.title = 'Segmentos CDA'
    if seg_cda_df is not None and len(seg_cda_df) > 0:
        write_seg_sheet(ws_cda, 'SEGMENTACAO — CDA / CUSUM (Zi)', seg_cda_df, fH2)
    else:
        ws_cda['A1'].value = 'Metodo CDA nao executado'

    # ── Aba Serie Zi ──
    ws_zi = wb.create_sheet('Serie Zi')
    if (col_zi_name is not None and picos_arr is not None and vales_arr is not None
            and col_zi_name in df_data.columns):
        hdrs = ['Estacao (m)', 'Zi', 'Zi (Picos)', 'Zi (Vales)']
        for j, h in enumerate(hdrs, 1):
            c = ws_zi.cell(row=1, column=j, value=h)
            c.font = fh; c.fill = fH2; c.alignment = ac; c.border = bd
        p_set = set(picos_arr); v_set = set(vales_arr)
        for i in range(len(df_data)):
            ws_zi.cell(row=2+i, column=1, value=float(df_data.loc[i, col_est])).number_format = '#,##0'
            ws_zi.cell(row=2+i, column=2, value=float(df_data.loc[i, col_zi_name])).number_format = '0.00'
            if i in p_set:
                ws_zi.cell(row=2+i, column=3, value=float(df_data.loc[i, col_zi_name]))
            if i in v_set:
                ws_zi.cell(row=2+i, column=4, value=float(df_data.loc[i, col_zi_name]))
        lz = 2 + len(df_data) - 1
        cz = LineChart(); cz.title = 'Serie Zi - Picos e Vales'
        cz.width = 32; cz.height = 16; cz.style = 10
        cz.y_axis.title = 'Zi'; cz.x_axis.title = 'Estacao (m)'
        cz.add_data(Reference(ws_zi, min_col=2, min_row=1, max_row=lz), titles_from_data=True)
        cz.add_data(Reference(ws_zi, min_col=3, min_row=1, max_row=lz), titles_from_data=True)
        cz.add_data(Reference(ws_zi, min_col=4, min_row=1, max_row=lz), titles_from_data=True)
        cz.set_categories(Reference(ws_zi, min_col=1, min_row=2, max_row=lz))
        cz.series[0].graphicalProperties.line.solidFill = COR_H[2:]
        cz.series[0].graphicalProperties.line.width = 22000
        cz.series[0].marker = Marker(symbol='none')
        cz.series[1].graphicalProperties.line.noFill = True
        cz.series[1].marker = Marker(symbol='triangle', size=10)
        cz.series[1].marker.graphicalProperties = GraphicalProperties()
        cz.series[1].marker.graphicalProperties.solidFill = COR_R[2:]
        cz.series[2].graphicalProperties.line.noFill = True
        cz.series[2].marker = Marker(symbol='triangle', size=10)
        cz.series[2].marker.graphicalProperties = GraphicalProperties()
        cz.series[2].marker.graphicalProperties.solidFill = COR_G[2:]
        if len(df_data) > 20:
            cz.x_axis.tickLblSkip = max(1, len(df_data) // 15)
        ws_zi.add_chart(cz, 'F1')
    else:
        ws_zi['A1'].value = 'Serie Zi nao disponivel (CDA nao executado)'

    # ── Aba SHS ──
    ws_shs = wb.create_sheet('Segmentos SHS')
    if seg_shs_df is not None and len(seg_shs_df) > 0:
        write_seg_sheet(ws_shs, 'SEGMENTACAO — Spatial Heterogeneity (SHS)', seg_shs_df, fAC)
    else:
        ws_shs['A1'].value = 'Metodo SHS nao executado'

    # ── Aba MCV ──
    ws_mcv = wb.create_sheet('Segmentos MCV')
    if seg_mcv_df is not None and len(seg_mcv_df) > 0:
        write_seg_sheet(ws_mcv, 'SEGMENTACAO — Minimize CV (MCV)', seg_mcv_df, fAC)
    else:
        ws_mcv['A1'].value = 'Metodo MCV nao executado'

    # ── Aba Dados Brutos ──
    ws_raw = wb.create_sheet('Dados Brutos')
    cols_export = [c for c in df_data.columns if c not in ['slk_from', 'slk_to']]
    for j, cn in enumerate(cols_export, 1):
        c = ws_raw.cell(row=1, column=j, value=cn)
        c.font = fh; c.fill = fH2; c.alignment = ac; c.border = bd
        ws_raw.column_dimensions[get_column_letter(j)].width = 20
    for i, rd in df_data.iterrows():
        for j, cn in enumerate(cols_export, 1):
            v = rd[cn]
            ws_raw.cell(row=2+i, column=j,
                        value=float(v) if isinstance(v, (int, float, np.integer, np.floating)) else v)

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ======================================================================
#  SIDEBAR
# ======================================================================
with st.sidebar:
    st.markdown("## \U0001f6e3\ufe0f Segmentacao Unificada")
    st.caption("CDA · SHS · MCV — 3 metodos em 1")
    st.divider()

    uploaded_file = st.file_uploader(
        "**Carregar planilha**",
        type=["xlsx", "xls", "csv"],
        help="Arquivo com colunas Estacao (m) e Deflexao (0,01mm)"
    )

    st.divider()
    st.markdown("### Colunas")
    col_estacao_name  = st.text_input("Coluna **Estacao**", value="Estação (m)")
    col_deflexao_name = st.text_input("Coluna **Deflexao**", value="Deflexão (0,01mm)")

    st.divider()
    st.markdown("### Metodo(s) de Segmentacao")
    metodos = st.multiselect(
        "Selecione os metodos a aplicar",
        options=["CDA (CUSUM/Zi)", "SHS (Spatial Heterogeneity)", "MCV (Minimize CV)"],
        default=["CDA (CUSUM/Zi)", "SHS (Spatial Heterogeneity)", "MCV (Minimize CV)"],
        help="Escolha um ou mais metodos."
    )

    st.divider()
    st.markdown("#### Restricoes de Comprimento")
    seg_min_m = st.number_input(
        "Comprimento MINIMO de SH (m)", value=800, min_value=50, step=50,
        help="Segmentos menores serao fundidos (CDA) ou restringidos (SHS/MCV).")
    seg_max_m = st.number_input(
        "Comprimento MAXIMO de SH (m)", value=5000, min_value=500, step=100,
        help="Segmentos maiores serao subdivididos (CDA) ou restringidos (SHS/MCV).")

    # ── Parametros CDA ──
    run_cda = "CDA (CUSUM/Zi)" in metodos
    if run_cda:
        st.divider()
        st.markdown("#### Parametros CDA (`find_peaks`)")
        fp_distance = st.slider("distance (pontos)", 1, 50, 5, 1,
                                help="Distancia minima entre picos consecutivos.")
        fp_height = st.number_input("height (Zi)", value=0.0, step=1.0, format="%.1f",
                                    help="Altura minima dos picos. 0=desativado.")
        fp_prominence = st.number_input("prominence (Zi)", value=0.0, min_value=0.0,
                                        step=1.0, format="%.1f")
        fp_width = st.number_input("width (pontos)", value=0.0, min_value=0.0,
                                   step=1.0, format="%.1f")
        fp_threshold = st.number_input("threshold (Zi)", value=0.0, min_value=0.0,
                                       step=0.5, format="%.1f")
        fp_plateau_size = st.number_input("plateau_size (pontos)", value=0, min_value=0, step=1)

    run_shs = "SHS (Spatial Heterogeneity)" in metodos
    run_mcv = "MCV (Minimize CV)" in metodos

    st.divider()
    st.caption(f"Gerado em {datetime.now():%d/%m/%Y %H:%M}")


# ======================================================================
#  HERO HEADER
# ======================================================================
st.markdown("""
<div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 30px 40px; border-radius: 16px; margin-bottom: 24px;
            box-shadow: 0 8px 30px rgba(102,126,234,0.35);">
    <h1 style="color: white; margin: 0; font-size: 2rem;">
        Segmentacao Homogenea de Rodovias
    </h1>
    <p style="color: rgba(255,255,255,0.85); font-size: 1.05rem; margin-top: 8px;">
        App unificado: <b>CDA</b> (Cumulative Difference Approach / CUSUM) ·
        <b>SHS</b> (Spatial Heterogeneity Segmentation) ·
        <b>MCV</b> (Minimize Coefficient of Variation)
    </p>
</div>
""", unsafe_allow_html=True)


# ======================================================================
#  PROCESSAMENTO
# ======================================================================
if uploaded_file is None:
    st.info("Carregue um arquivo Excel ou CSV na barra lateral para iniciar a analise.")
    st.markdown("""
    ### O que este app faz?
    1. **Le** uma planilha com colunas **Estacao** e **Deflexao**
    2. **Aplica** ate 3 metodos de segmentacao homogenea
    3. **Gera** visualizacoes interativas (Plotly) e exporta planilha Excel formatada

    ---
    **Metodos disponiveis:**

    | Metodo | Descricao |
    |--------|-----------|
    | **CDA** | Diferenca acumulada (Zi / CUSUM) com deteccao de picos/vales via `scipy` |
    | **SHS** | Maximiza a heterogeneidade espacial entre segmentos |
    | **MCV** | Minimiza o coeficiente de variacao dentro de cada segmento |
    """)
    st.stop()

if not metodos:
    st.warning("Selecione ao menos um metodo de segmentacao na barra lateral.")
    st.stop()

# -- Leitura dos dados --
fname = uploaded_file.name
try:
    if fname.endswith('.csv'):
        df = pd.read_csv(uploaded_file, sep=None, engine='python')
    else:
        df = pd.read_excel(uploaded_file)
except Exception as e:
    st.error(f"Erro ao ler o arquivo: {e}")
    st.stop()

col_est  = col_estacao_name
col_defl = col_deflexao_name

if col_est not in df.columns or col_defl not in df.columns:
    st.error(f"Colunas esperadas nao encontradas. Disponiveis: {list(df.columns)}")
    st.stop()

for c in [col_est, col_defl]:
    if df[c].dtype == 'O':
        df[c] = df[c].astype(str).str.replace(',', '.').astype(float)
    df[c] = pd.to_numeric(df[c], errors='coerce')
df.dropna(subset=[col_est, col_defl], inplace=True)
df.sort_values(col_est, inplace=True)
df.reset_index(drop=True, inplace=True)

# -- Filtro de Estacao na sidebar --
with st.sidebar:
    st.divider()
    st.markdown("#### Filtro de Estacao")
    est_min_data = float(df[col_est].min())
    est_max_data = float(df[col_est].max())
    est_range = st.slider(
        "Intervalo de Estacao (m)",
        min_value=est_min_data, max_value=est_max_data,
        value=(est_min_data, est_max_data),
        step=10.0, format="%.0f",
        help="Selecione o trecho da rodovia a ser analisado."
    )
    est_filter_min, est_filter_max = est_range

df = df[(df[col_est] >= est_filter_min) & (df[col_est] <= est_filter_max)].copy()
df.reset_index(drop=True, inplace=True)

if len(df) < 3:
    st.error("Poucos registros apos o filtro de Estacao. Amplie o intervalo.")
    st.stop()

media_deflexao = df[col_defl].mean()


# ======================================================================
#  EXECUTAR METODOS
# ======================================================================

# ── CDA ──
seg_cda_df = None
picos = np.array([]); vales = np.array([])
limites_idx = np.array([]); picos_raw = np.array([]); vales_raw = np.array([])
col_diff = 'Diferenca'; col_zi = 'Diferenca Acumulada (Zi)'
zi = np.array([])
fp_kwargs = {}

if run_cda:
    df[col_diff] = df[col_defl] - media_deflexao
    df[col_zi]   = df[col_diff].cumsum()
    zi = df[col_zi].values

    fp_kwargs = {'distance': max(1, fp_distance)}
    if fp_height > 0:       fp_kwargs['height'] = fp_height
    if fp_prominence > 0:   fp_kwargs['prominence'] = fp_prominence
    if fp_width > 0:        fp_kwargs['width'] = fp_width
    if fp_threshold > 0:    fp_kwargs['threshold'] = fp_threshold
    if fp_plateau_size > 0: fp_kwargs['plateau_size'] = fp_plateau_size

    picos_raw, _ = find_peaks(zi, **fp_kwargs)
    vales_raw, _ = find_peaks(-zi, **fp_kwargs)
    limites_ini = np.sort(np.concatenate(([0], picos_raw, vales_raw, [len(df)-1])))

    limites_merged = merge_short_segments(limites_ini, df, col_est, seg_min_m)
    limites_idx    = split_long_segments(limites_merged, df, col_est, seg_max_m)

    picos = np.array(sorted(set(picos_raw) & set(limites_idx)))
    vales = np.array(sorted(set(vales_raw) & set(limites_idx)))

    segmentos = []
    for i in range(len(limites_idx) - 1):
        ii, fi = limites_idx[i], limites_idx[i+1]
        est_i = float(df.loc[ii, col_est]); est_f = float(df.loc[fi, col_est])
        defl_seg = df.loc[ii:fi, col_defl]
        defl_media = defl_seg.mean()
        defl_std   = defl_seg.std(ddof=1) if len(defl_seg) > 1 else 0.0
        cv = (defl_std / defl_media * 100) if defl_media != 0 else 0.0
        segmentos.append({
            'SH': i+1,
            'Inicio (m)': est_i,  'Fim (m)': est_f,
            'Comprimento (m)': est_f - est_i,
            'Zi_fim': round(float(df.loc[fi, col_zi]), 1),
            'Defl. Media': round(defl_media, 1),
            'Desvio Padrao': round(defl_std, 1),
            'CV (%)': round(cv, 1),
            'Defl. Caract.': round(defl_media + defl_std, 1),
        })
    seg_cda_df = pd.DataFrame(segmentos)


# ── SHS / MCV ──
seg_shs_df = None
seg_mcv_df = None

if run_shs or run_mcv:
    step = df[col_est].diff().median()
    df['slk_from'] = df[col_est]
    df['slk_to'] = df[col_est].shift(-1).fillna(df[col_est].iloc[-1] + step)
    seg_range = (float(seg_min_m), float(seg_max_m))

if run_shs:
    try:
        df['seg_shs'] = segment_ids_to_maximize_spatial_heterogeneity(
            data=df, measure=("slk_from", "slk_to"),
            variable_column_names=[col_defl],
            allowed_segment_length_range=seg_range,
        )
        seg_shs_df = build_segments_table(df, 'seg_shs', col_est, col_defl)
    except Exception as e:
        st.error(f"Erro no metodo SHS: {e}")

if run_mcv:
    try:
        df['seg_mcv'] = segment_ids_to_minimize_coefficient_of_variation(
            data=df, measure=("slk_from", "slk_to"),
            variable_column_names=[col_defl],
            allowed_segment_length_range=seg_range,
        )
        seg_mcv_df = build_segments_table(df, 'seg_mcv', col_est, col_defl)
    except Exception as e:
        st.error(f"Erro no metodo MCV: {e}")


# Verificar que ao menos 1 metodo retornou resultado
results = {'CDA': seg_cda_df, 'SHS': seg_shs_df, 'MCV': seg_mcv_df}
active_results = {k: v for k, v in results.items() if v is not None and len(v) > 0}

if not active_results:
    st.error("Nenhum metodo retornou resultados. Verifique os parametros.")
    st.stop()


# ======================================================================
#  FUNCAO DE RENDERIZACAO GENERICA POR METODO
# ======================================================================

def render_method_tab(method_key, seg_df, extra_info=""):
    """Renderiza KPIs, graficos e tabela para um metodo qualquer."""

    st.markdown(f"### Indicadores — {method_key}")

    cols_kpi = st.columns(5)
    with cols_kpi[0]:
        st.markdown(kpi_card(f"{len(seg_df)}", "Segmentos"), unsafe_allow_html=True)
    with cols_kpi[1]:
        st.markdown(kpi_card(f"{seg_df['Comprimento (m)'].sum():,.0f} m",
                             "Extensao Total", "kpi-green"), unsafe_allow_html=True)
    with cols_kpi[2]:
        st.markdown(kpi_card(f"{seg_df['Comprimento (m)'].mean():,.0f} m",
                             "Compr. Medio", "kpi-blue"), unsafe_allow_html=True)
    with cols_kpi[3]:
        st.markdown(kpi_card(f"{seg_df['CV (%)'].mean():.1f}%",
                             "CV Medio", "kpi-purple"), unsafe_allow_html=True)
    with cols_kpi[4]:
        st.markdown(kpi_card(f"{media_deflexao:.1f}",
                             "Defl. Media Global", "kpi-orange"), unsafe_allow_html=True)

    st.markdown("")
    story(
        f"O metodo <b>{method_key}</b> identificou <b>{len(seg_df)} segmentos homogeneos</b> "
        f"ao longo de <b>{seg_df['Comprimento (m)'].sum():,.0f} m</b>. "
        f"O CV medio dos segmentos e <b>{seg_df['CV (%)'].mean():.1f}%</b>. "
        f"Restricoes: [{seg_min_m} m, {seg_max_m} m]. {extra_info}"
    )
    st.divider()

    # -- Deflexao ao longo da rodovia --
    st.markdown(f"#### Deflexao ao Longo da Rodovia — {method_key}")
    fig_defl = go.Figure()
    for i, row in seg_df.iterrows():
        mask = (df[col_est] >= row['Inicio (m)']) & (df[col_est] <= row['Fim (m)'])
        hex_c = SEG_COLORS_HEX[i % len(SEG_COLORS_HEX)]
        fig_defl.add_trace(go.Bar(
            x=df.loc[mask, col_est], y=df.loc[mask, col_defl],
            marker_color=hex_c, opacity=0.7,
            name=f"SH {row['SH']} ({row['Comprimento (m)']:,.0f}m)",
            legendgroup=f"sh{row['SH']}",
        ))
    fig_defl.add_hline(y=media_deflexao, line_dash='dash', line_color=COLOR_RED, line_width=2,
                       annotation_text=f'Media = {media_deflexao:.2f}',
                       annotation_position='top right')
    fig_defl.update_layout(
        xaxis_title='Estacao (m)', yaxis_title='Deflexao (0,01mm)',
        template='plotly_white', height=420, margin=dict(t=30, b=60),
        legend=dict(orientation='h', y=-0.18, xanchor='center', x=0.5, font_size=10),
        barmode='stack',
    )
    st.plotly_chart(fig_defl, use_container_width=True)

    # -- Deflexao media por segmento --
    st.markdown(f"#### Deflexao Media por Segmento — {method_key}")
    fig_bar = go.Figure()
    bar_colors = [COLOR_RED if d > seg_df['Defl. Media'].mean() * 1.2
                  else COLOR_YELLOW if d > seg_df['Defl. Media'].mean()
                  else COLOR_GREEN for d in seg_df['Defl. Media']]
    fig_bar.add_trace(go.Bar(
        x=seg_df['SH'].apply(lambda x: f'SH {x}'), y=seg_df['Defl. Media'],
        marker_color=bar_colors, showlegend=False,
        text=seg_df['Defl. Media'].apply(lambda x: f'{x:.1f}'), textposition='outside',
    ))
    fig_bar.add_hline(y=seg_df['Defl. Media'].mean(), line_dash='dash',
                      line_color=COLOR_PRIMARY, line_width=2,
                      annotation_text=f"Media = {seg_df['Defl. Media'].mean():.1f}")
    fig_bar.update_layout(template='plotly_white', height=380, margin=dict(t=20, b=40),
                          xaxis_title='Segmento', yaxis_title='Defl. Media (0,01mm)')
    st.plotly_chart(fig_bar, use_container_width=True)

    # -- CV por segmento --
    st.markdown(f"#### Coeficiente de Variacao por Segmento — {method_key}")
    story("CV = Desvio Padrao / Media x 100. Valores <= 25% sao considerados homogeneos (Austroads AGPT05-19).")
    fig_cv = go.Figure()
    cv_colors = [COLOR_GREEN if cv <= 25 else COLOR_YELLOW if cv <= 40 else COLOR_RED
                 for cv in seg_df['CV (%)']]
    fig_cv.add_trace(go.Bar(
        x=seg_df['SH'].apply(lambda x: f'SH {x}'), y=seg_df['CV (%)'],
        marker_color=cv_colors, showlegend=False,
        text=seg_df['CV (%)'].apply(lambda x: f'{x:.1f}%'), textposition='outside',
    ))
    fig_cv.add_hline(y=25, line_dash='dash', line_color=COLOR_GREEN, line_width=2,
                     annotation_text="Limite Austroads (25%)")
    fig_cv.update_layout(template='plotly_white', height=380, margin=dict(t=20, b=40),
                         xaxis_title='Segmento', yaxis_title='CV (%)')
    st.plotly_chart(fig_cv, use_container_width=True)

    # -- Boxplot --
    st.markdown(f"#### Variabilidade da Deflexao por Segmento — {method_key}")
    box_data = []
    for _, row in seg_df.iterrows():
        mask = (df[col_est] >= row['Inicio (m)']) & (df[col_est] <= row['Fim (m)'])
        for v in df.loc[mask, col_defl].values:
            box_data.append({'Segmento': f"SH {int(row['SH'])}", 'Deflexao': v})
    df_box = pd.DataFrame(box_data)
    fig_box = px.box(df_box, x='Segmento', y='Deflexao', color='Segmento',
                     color_discrete_sequence=px.colors.qualitative.Set2)
    fig_box.add_hline(y=media_deflexao, line_dash='dash', line_color=COLOR_RED,
                      annotation_text=f'Media global = {media_deflexao:.1f}')
    fig_box.update_layout(template='plotly_white', height=400,
                          yaxis_title='Deflexao (0,01mm)', showlegend=False, margin=dict(t=30))
    st.plotly_chart(fig_box, use_container_width=True)

    # -- Tabela --
    st.markdown(f"#### Tabela de Segmentos — {method_key}")
    n_curtos = (seg_df['Comprimento (m)'] < seg_min_m).sum()
    n_longos = (seg_df['Comprimento (m)'] > seg_max_m).sum()
    if n_curtos == 0 and n_longos == 0:
        st.success(f"Todos os {len(seg_df)} segmentos atendem as restricoes "
                   f"({seg_min_m}m <= SH <= {seg_max_m}m)")
    else:
        if n_curtos: st.warning(f"{n_curtos} segmento(s) abaixo de {seg_min_m}m")
        if n_longos: st.warning(f"{n_longos} segmento(s) acima de {seg_max_m}m")

    format_dict = {'Inicio (m)': '{:,.0f}', 'Fim (m)': '{:,.0f}',
                   'Comprimento (m)': '{:,.0f}', 'CV (%)': '{:.1f}',
                   'Defl. Media': '{:.1f}', 'Desvio Padrao': '{:.1f}',
                   'Defl. Caract.': '{:.1f}'}
    if 'Zi_fim' in seg_df.columns:
        format_dict['Zi_fim'] = '{:.1f}'

    st.dataframe(
        seg_df.style
            .background_gradient(subset=['Comprimento (m)'], cmap='Blues')
            .background_gradient(subset=['Defl. Media'], cmap='RdYlGn_r')
            .background_gradient(subset=['CV (%)'], cmap='YlOrRd')
            .background_gradient(subset=['Defl. Caract.'], cmap='OrRd')
            .format(format_dict),
        height=400
    )

    st.markdown("##### Resumo Estatistico")
    desc_cols = ['Comprimento (m)', 'Defl. Media', 'CV (%)']
    if 'Zi_fim' in seg_df.columns:
        desc_cols.append('Zi_fim')
    st.dataframe(seg_df[desc_cols].describe().round(2))


# ======================================================================
#  ABAS PRINCIPAIS
# ======================================================================
tab_labels = []
tab_keys   = []

for key in ['CDA', 'SHS', 'MCV']:
    if key in active_results:
        tab_labels.append(key)
        tab_keys.append(key)

if run_cda and seg_cda_df is not None:
    tab_labels.append("Serie Zi")
    tab_keys.append("__zi__")

if len(active_results) >= 2:
    tab_labels.append("Comparacao")
    tab_keys.append("__comp__")

tab_labels += ["Dados", "Exportar"]
tab_keys   += ["__dados__", "__export__"]

all_tabs = st.tabs(tab_labels)
tab_map = dict(zip(tab_keys, all_tabs))


# -- TABS POR METODO --
for key in ['CDA', 'SHS', 'MCV']:
    if key not in tab_map:
        continue
    with tab_map[key]:
        seg = active_results[key]
        extra = ""
        if key == 'CDA':
            extra = (f"Picos brutos: {len(picos_raw)}, Vales brutos: {len(vales_raw)}, "
                     f"Limites finais: {len(limites_idx)}.")
        render_method_tab(key, seg, extra)


# -- TAB SERIE ZI (CDA only) --
if "__zi__" in tab_map and seg_cda_df is not None:
    with tab_map["__zi__"]:
        st.markdown("#### Deflexao ao Longo da Rodovia")
        story("Grafico de barras da deflexao medida. "
              "A linha tracejada vermelha indica a media global.")

        fig_defl_zi = go.Figure()
        fig_defl_zi.add_trace(go.Bar(
            x=df[col_est], y=df[col_defl], name='Deflexao',
            marker_color=COLOR_PRIMARY, opacity=0.6,
        ))
        fig_defl_zi.add_hline(y=media_deflexao, line_dash='dash', line_color=COLOR_RED, line_width=2,
                              annotation_text=f'Media = {media_deflexao:.2f}',
                              annotation_position='top right')
        fig_defl_zi.update_layout(xaxis_title='Estacao (m)', yaxis_title='Deflexao (0,01mm)',
                                  template='plotly_white', height=380, margin=dict(t=30, b=40),
                                  legend=dict(orientation='h', y=-0.15))
        st.plotly_chart(fig_defl_zi, use_container_width=True)

        st.markdown("#### Serie de Diferencas Acumuladas (Zi)")
        story("Curva Zi = soma cumulativa de (Deflexao - Media). "
              "Picos e vales marcam mudancas de comportamento que definem limites de SH.")

        fig_zi = go.Figure()
        for i, row in seg_cda_df.iterrows():
            mask = (df[col_est] >= row['Inicio (m)']) & (df[col_est] <= row['Fim (m)'])
            hex_c = SEG_COLORS_HEX[i % len(SEG_COLORS_HEX)]
            fig_zi.add_trace(go.Scatter(
                x=df.loc[mask, col_est], y=zi[mask],
                fill='tozeroy', mode='lines', line=dict(width=0),
                fillcolor=rgba_from_hex(hex_c, 0.2),
                name=f"SH {row['SH']}", showlegend=False, hoverinfo='skip'))
        fig_zi.add_trace(go.Scatter(
            x=df[col_est], y=zi, mode='lines', name='Zi',
            line=dict(color='#2c3e50', width=2.5)))
        if len(picos) > 0:
            fig_zi.add_trace(go.Scatter(
                x=df.loc[picos, col_est].values, y=zi[picos],
                mode='markers+text', name=f'Picos ({len(picos)})',
                marker=dict(symbol='triangle-up', size=14, color=COLOR_RED,
                            line=dict(color='white', width=2)),
                text=[f'{v:.0f}' for v in zi[picos]],
                textposition='top center', textfont=dict(size=9, color=COLOR_RED)))
        if len(vales) > 0:
            fig_zi.add_trace(go.Scatter(
                x=df.loc[vales, col_est].values, y=zi[vales],
                mode='markers+text', name=f'Vales ({len(vales)})',
                marker=dict(symbol='triangle-down', size=14, color=COLOR_GREEN,
                            line=dict(color='white', width=2)),
                text=[f'{v:.0f}' for v in zi[vales]],
                textposition='bottom center', textfont=dict(size=9, color=COLOR_GREEN)))
        for idx in limites_idx[1:-1]:
            fig_zi.add_vline(x=float(df.loc[idx, col_est]), line_dash='dot',
                             line_color='gray', line_width=0.8, opacity=0.5)
        fig_zi.add_hline(y=0, line_dash='dash', line_color='gray', line_width=1, opacity=0.5)
        fig_zi.update_layout(
            xaxis_title='Estacao (m)', yaxis_title='Diferenca Acumulada (Zi)',
            template='plotly_white', height=500, margin=dict(t=30, b=40),
            legend=dict(orientation='h', y=-0.12), hovermode='x unified')
        st.plotly_chart(fig_zi, use_container_width=True)

        # Mapa de segmentos coloridos
        st.markdown("#### Mapa de Segmentos — Serie Zi Colorida")
        fig_zi_seg = go.Figure()
        for i, row in seg_cda_df.iterrows():
            mask = (df[col_est] >= row['Inicio (m)']) & (df[col_est] <= row['Fim (m)'])
            hex_c = SEG_COLORS_HEX[i % len(SEG_COLORS_HEX)]
            fig_zi_seg.add_trace(go.Scatter(
                x=df.loc[mask, col_est], y=zi[mask],
                fill='tozeroy', mode='lines', line=dict(width=0.5, color=hex_c),
                fillcolor=rgba_from_hex(hex_c, 0.45),
                name=f"SH {row['SH']} ({row['Comprimento (m)']:,.0f}m)",
                legendgroup=f"sh{row['SH']}", showlegend=True,
            ))
        fig_zi_seg.add_trace(go.Scatter(
            x=df[col_est], y=zi, mode='lines', name='Zi',
            line=dict(color='#2c3e50', width=2), showlegend=False,
        ))
        fig_zi_seg.update_layout(
            template='plotly_white', height=450, margin=dict(t=20, b=60),
            xaxis_title='Estacao (m)', yaxis_title='Zi',
            legend=dict(orientation='h', y=-0.18, font_size=10, xanchor='center', x=0.5),
            hovermode='x unified',
        )
        st.plotly_chart(fig_zi_seg, use_container_width=True)


# -- TAB COMPARACAO --
if "__comp__" in tab_map:
    with tab_map["__comp__"]:
        st.markdown("#### Comparacao entre Metodos")
        story("Comparacao lado a lado dos metodos de segmentacao ativados.")

        # KPIs lado a lado
        comp_cols = st.columns(len(active_results))
        method_colors = {'CDA': 'kpi-blue', 'SHS': 'kpi-purple', 'MCV': 'kpi-teal'}
        for col_ui, (mkey, mdf) in zip(comp_cols, active_results.items()):
            with col_ui:
                st.markdown(f"##### {mkey}")
                st.markdown(kpi_card(f"{len(mdf)}", f"Segmentos {mkey}",
                                     method_colors.get(mkey, '')), unsafe_allow_html=True)
                st.markdown(kpi_card(f"{mdf['CV (%)'].mean():.1f}%", "CV Medio", "kpi-orange"),
                            unsafe_allow_html=True)
                st.dataframe(
                    mdf[['SH','Inicio (m)','Fim (m)','Comprimento (m)',
                         'Defl. Media','CV (%)']].style.format(
                        {'Inicio (m)': '{:,.0f}', 'Fim (m)': '{:,.0f}',
                         'Comprimento (m)': '{:,.0f}', 'CV (%)': '{:.1f}',
                         'Defl. Media': '{:.1f}'}), height=350)

        st.divider()

        # Grafico comparativo: deflexao media
        st.markdown("##### Deflexao Media — Comparacao")
        fig_comp = go.Figure()
        method_plotcolors = {'CDA': COLOR_PRIMARY, 'SHS': COLOR_SEC, 'MCV': COLOR_TEAL}
        for mkey, mdf in active_results.items():
            fig_comp.add_trace(go.Bar(
                x=mdf['SH'].apply(lambda x: f'SH {x}'),
                y=mdf['Defl. Media'], name=mkey,
                marker_color=method_plotcolors.get(mkey, COLOR_PRIMARY), opacity=0.8,
                text=mdf['Defl. Media'].apply(lambda x: f'{x:.1f}'), textposition='outside',
            ))
        fig_comp.update_layout(barmode='group', template='plotly_white', height=400,
                               xaxis_title='Segmento', yaxis_title='Defl. Media (0,01mm)',
                               legend=dict(orientation='h', y=-0.15))
        st.plotly_chart(fig_comp, use_container_width=True)

        # Grafico comparativo: CV
        st.markdown("##### CV (%) — Comparacao")
        fig_cv_comp = go.Figure()
        for mkey, mdf in active_results.items():
            fig_cv_comp.add_trace(go.Bar(
                x=mdf['SH'].apply(lambda x: f'SH {x}'),
                y=mdf['CV (%)'], name=mkey,
                marker_color=method_plotcolors.get(mkey, COLOR_PRIMARY), opacity=0.8,
            ))
        fig_cv_comp.add_hline(y=25, line_dash='dash', line_color=COLOR_GREEN,
                              annotation_text="Limite Austroads (25%)")
        fig_cv_comp.update_layout(barmode='group', template='plotly_white', height=400,
                                  xaxis_title='Segmento', yaxis_title='CV (%)',
                                  legend=dict(orientation='h', y=-0.15))
        st.plotly_chart(fig_cv_comp, use_container_width=True)

        # Resumo numerico
        st.markdown("##### Resumo Numerico Comparativo")
        resumo_data = []
        for mkey, mdf in active_results.items():
            resumo_data.append({
                'Metodo': mkey,
                'Segmentos': len(mdf),
                'Extensao Total (m)': f"{mdf['Comprimento (m)'].sum():,.0f}",
                'Compr. Medio (m)': f"{mdf['Comprimento (m)'].mean():,.0f}",
                'CV Medio (%)': f"{mdf['CV (%)'].mean():.1f}",
                'Defl. Media': f"{mdf['Defl. Media'].mean():.1f}",
                'Homogeneos (CV<=25%)': f"{(mdf['CV (%)'] <= 25).sum()}/{len(mdf)}",
            })
        st.dataframe(pd.DataFrame(resumo_data), hide_index=True)


# -- TAB DADOS --
with tab_map["__dados__"]:
    st.markdown("#### Dados Brutos Processados")
    cols_show = [c for c in df.columns if c not in ['slk_from', 'slk_to']]
    st.caption(f"{len(df)} registros")
    st.dataframe(df[cols_show], height=450)


# -- TAB EXPORTAR --
with tab_map["__export__"]:
    st.markdown("#### Exportar Resultados")
    story("Baixe os resultados em <b>Excel</b> (com graficos e formatacao) ou <b>CSV</b>.")

    col_e1, col_e2 = st.columns(2)

    with col_e1:
        abas_desc = ", ".join([f"Seg. {k}" for k in active_results.keys()])
        if run_cda and seg_cda_df is not None:
            abas_desc += ", Serie Zi"
        st.markdown(f"""
        <div style="background:#f0f2f6; padding:20px; border-radius:12px; text-align:center;">
            <div style="font-size:3rem;">&#128215;</div>
            <div style="font-weight:bold; margin:8px 0;">Planilha Excel</div>
            <div style="font-size:0.85rem; color:#666;">
                Abas: {abas_desc}, Dados Brutos
            </div>
        </div>
        """, unsafe_allow_html=True)
        excel_bytes = gerar_excel_unificado(
            df, seg_cda_df, seg_shs_df, seg_mcv_df,
            col_est, col_defl,
            picos_arr=picos, vales_arr=vales,
            limites_idx_arr=limites_idx,
            col_zi_name=col_zi if run_cda else None)
        st.download_button(
            "Baixar Excel",
            data=excel_bytes,
            file_name=f"segmentos_unificado_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    with col_e2:
        st.markdown("""
        <div style="background:#f0f2f6; padding:20px; border-radius:12px; text-align:center;">
            <div style="font-size:3rem;">&#128196;</div>
            <div style="font-weight:bold; margin:8px 0;">CSV (Segmentos)</div>
            <div style="font-size:0.85rem; color:#666;">
                Todas as tabelas de segmentos<br>
                Separador: ponto e virgula (;)
            </div>
        </div>
        """, unsafe_allow_html=True)
        csv_parts = []
        for mkey, mdf in active_results.items():
            csv_parts.append(f"# METODO: {mkey}\n" + mdf.to_csv(index=False, sep=';'))
        csv_data = "\n\n".join(csv_parts).encode('utf-8-sig')
        st.download_button(
            "Baixar CSV",
            data=csv_data,
            file_name=f"segmentos_unificado_{datetime.now():%Y%m%d_%H%M%S}.csv",
            mime="text/csv",
        )

    st.divider()
    st.markdown("#### Parametros Utilizados")
    params_rows = [
        {'Parametro': 'Metodo(s)', 'Valor': ', '.join(active_results.keys())},
        {'Parametro': 'seg_min_m', 'Valor': f'{seg_min_m} m'},
        {'Parametro': 'seg_max_m', 'Valor': f'{seg_max_m} m'},
        {'Parametro': 'Media deflexao', 'Valor': f'{media_deflexao:.2f}'},
        {'Parametro': 'Registros', 'Valor': str(len(df))},
    ]
    if run_cda:
        fp_desc = ', '.join(f'{k}={v}' for k, v in fp_kwargs.items())
        params_rows.append({'Parametro': 'find_peaks kwargs', 'Valor': fp_desc})
        params_rows.append({'Parametro': 'Picos (brutos)', 'Valor': str(len(picos_raw))})
        params_rows.append({'Parametro': 'Vales (brutos)', 'Valor': str(len(vales_raw))})
        params_rows.append({'Parametro': 'Limites finais CDA', 'Valor': str(len(limites_idx))})
    for mkey in ['CDA', 'SHS', 'MCV']:
        if mkey in active_results:
            params_rows.append({'Parametro': f'Segmentos {mkey}',
                                'Valor': str(len(active_results[mkey]))})
    st.dataframe(pd.DataFrame(params_rows), hide_index=True)


# ======================================================================
#  FOOTER
# ======================================================================
st.divider()
st.markdown("""
<div style="text-align: center; color: #999; font-size: 0.8rem; padding: 10px;">
    Segmentacao Homogenea Unificada — CDA · SHS · MCV | Streamlit + Plotly
</div>
""", unsafe_allow_html=True)

